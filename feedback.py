import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# --- 定数定義：Excelのフォーマットに合わせてここを調整してください ---
TARGET_SHEET_NAME = "サービスコードマスタ（入力シート）"  # 操作対象のワークシート名
HEADER_ROW_INDEX = 2  # ヘッダー（列名）がある行番号 (1開始)
DATA_START_ROW_INDEX = 7  # データが開始される行番号 (1開始)


def mark_and_update_excel_errors(file_path, output_path, error_json, pk_columns):
    """
    エラー箇所の特定、値の書き換え（異常値）、ハイライト、およびコメント（判定理由・修正アドバイス）の追加を行う。
    """
    try:
        # 1. ワークブックの読み込み (openpyxl)
        wb = load_workbook(file_path)
        if TARGET_SHEET_NAME in wb.sheetnames:
            ws = wb[TARGET_SHEET_NAME]
        else:
            print(f"エラー: シート名 '{TARGET_SHEET_NAME}' が見つかりません。")
            return

        # 2. データの読み込み (照合用：pandas)
        df_full = pd.read_excel(file_path, sheet_name=TARGET_SHEET_NAME, header=HEADER_ROW_INDEX - 1, dtype=str)

        # 列名の改行コードを除去して正規化
        df_full.columns = [str(col).replace('\n', '').strip() for col in df_full.columns]

        # データ開始行に合わせてデータフレームを抽出
        data_offset = DATA_START_ROW_INDEX - HEADER_ROW_INDEX - 1
        df = df_full.iloc[data_offset:].reset_index(drop=True)

        # 3. 複合主キー -> Excel行番号のマッピング作成
        pk_to_row = {}
        for idx, row in df.iterrows():
            composite_key = tuple(str(row.get(col, "")).strip() for col in pk_columns)
            pk_to_row[composite_key] = DATA_START_ROW_INDEX + idx

        # 4. 列名 -> 列インデックス(1, 2...)のマッピング作成
        col_to_idx = {}
        for cell in ws[HEADER_ROW_INDEX]:
            if cell.value:
                clean_label = str(cell.value).replace('\n', '').strip()
                col_to_idx[clean_label] = cell.column

        # セルの塗りつぶしスタイル（赤色）
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        # 5. エラーデータに基づきセルを更新
        for error in error_json:
            # 主キーの値を抽出してタプル化
            current_pk_values = tuple(str(error.get(col)).strip() for col in pk_columns)

            # 指定された新しいカラム名で値を取得
            target_col = str(error.get("異常フィールド", "")).replace('\n', '').strip()
            new_value = error.get("異常値")
            reason = error.get("判定理由", "エラーあり")
            suggestion = error.get("修正アドバイス", "内容を確認してください")

            # 行と列がマッピングに存在する場合のみ処理を実行
            if current_pk_values in pk_to_row and target_col in col_to_idx:
                row_idx = pk_to_row[current_pk_values]
                col_idx = col_to_idx[target_col]
                cell = ws.cell(row=row_idx, column=col_idx)

                # --- 値の書き換え（異常値に置換） ---
                # if pd.notna(new_value):
                #     cell.value = str(new_value)

                # --- 装飾：背景色とコメントの追加 ---
                cell.fill = red_fill
                comment_text = f"【判定理由】: {reason}\n【修正アドバイス】: {suggestion}"
                cell.comment = Comment(comment_text, author="AI監査システム")
            else:
                print(f"警告: 行({current_pk_values}) または 列({target_col}) が見つかりませんでした。")

        # 6. 保存処理
        wb.save(output_path)
        print(f"完了: 修正結果を '{output_path}' に保存しました。")

    except Exception as e:
        print(f"実行中に予期せぬエラーが発生しました: {e}")


# --- メイン処理 ---
if __name__ == "__main__":
    # 外部のエラーリストファイルを指定
    error_list_file = "error_input.xlsx"

    try:
        # エラーリストを読み込み、辞書形式のリストに変換
        error_df = pd.read_excel(error_list_file, dtype=str)
        error_data_json = error_df.to_dict(orient='records')

        # 主キーとして使用するカラム名
        pks = ["プロジェクトコード", "枝番"]

        # 実行
        mark_and_update_excel_errors("data_dual.xlsx", "data_fixed.xlsx", error_data_json, pks)
    except Exception as e:
        print(f"エラーリストの読み込みに失敗しました: {e}")