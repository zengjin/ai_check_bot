import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import traceback
import re
import warnings

# 忽略 openpyxl 产生的特定 UserWarning
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- 定数定義：Excelのフォーマットに合わせてここを調整してください ---
TARGET_SHEET_NAME = "サービスコードマスタ（入力シート）"  # 操作対象のワークシート名
HEADER_ROW_INDEX = 2  # ヘッダー（列名）がある行番号 (1開始)
DATA_START_ROW_INDEX = 7  # データが開始される行番号 (1開始)

# --- デバッグスイッチ ---
DEBUG_MODE = True  # True: 詳細なログを出力, False: 最終結果のみ出力

def log(message):
    """デバッグモードが有効な場合のみメッセージを表示する"""
    if DEBUG_MODE:
        print(message)

def mark_and_update_excel_errors(file_path, output_path, error_json, pk_columns):
    """
    エラー箇所の特定、ハイライト、およびコメント（判定理由・修正アドバイス）の追加を行う。
    """
    try:
        log(f"--- [START] 処理開始: {file_path} ---")
        
        # 1. ワークブックの読み込み (openpyxl)
        wb = load_workbook(file_path, data_only=True)  # data_only=Trueで数式ではなく値を取得
        if TARGET_SHEET_NAME in wb.sheetnames:
            ws = wb[TARGET_SHEET_NAME]
            log(f"[OK] シート確認: {TARGET_SHEET_NAME}")
        else:
            print(f"[ERROR] シート名 '{TARGET_SHEET_NAME}' が見つかりません。")
            return

        # 2. データの読み込み (照合用：pandas)
        df_full = pd.read_excel(file_path, sheet_name=TARGET_SHEET_NAME, header=HEADER_ROW_INDEX - 1, dtype=str, keep_default_na=False)

        # 列名の改行コードを除去して正規化
        df_full.columns = [str(col).replace('\n', '').strip() for col in df_full.columns]
        log(f"[DEBUG] 読み込まれた列名: {list(df_full.columns)}")

        # データ開始行に合わせてデータフレームを抽出
        data_offset = DATA_START_ROW_INDEX - HEADER_ROW_INDEX - 1
        df = df_full.iloc[data_offset:].reset_index(drop=True)
        log(f"[DEBUG] データオフセット: {data_offset}, 処理対象データ件数: {len(df)}")

        # 3. 複合主キー -> Excel行番号のマッピング作成
        pk_to_row = {}
        for idx, row in df.iterrows():
            composite_key = tuple(str(row.get(col, "")).strip() for col in pk_columns)
            pk_to_row[composite_key] = DATA_START_ROW_INDEX + idx
        
        log(f"[DEBUG] 主キーマッピング作成完了 ({len(pk_to_row)} 件)")

        # 4. 列名 -> 列インデックス(1, 2...)のマッピング作成
        col_to_idx = {}
        for cell in ws[HEADER_ROW_INDEX]:
            if cell.value:
                clean_label = str(cell.value).replace('\n', '').strip()
                col_to_idx[clean_label] = cell.column
        
        # セルの塗りつぶしスタイル（赤色）
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        # 5. エラーデータに基づきセルを更新
        success_count = 0
        fail_count = 0
        
        log(f"--- [PROCESS] エラー照合中 (全 {len(error_json)} 件) ---")
        
        for error in error_json:
            # 主キーの値を抽出してタプル化
            current_pk_values = tuple(str(error.get(col)).strip() for col in pk_columns)

            # # --- 複数フィールドの分割対応 ---
            # raw_target_cols = str(error.get("異常フィールド", ""))
            # # カンマ（全角・半角）で分割し、前後の空白を除去
            # target_cols = [c.strip() for re_split in [re.split(r'[,]', raw_target_cols)] for c in re_split if c.strip()]
            
            # --- 複数フィールドの分割対応（修正版） ---
            raw_target_cols = str(error.get("異常フィールド", ""))
            
            # カンマ(,) または スラッシュ(/) または セミコロン(;) で分割
            # [,/;/] のように[]内に区切り文字を追加します
            target_cols = [c.strip() for c in re.split(r'[,/;/]', raw_target_cols) if c.strip()]

            reason = error.get("判定理由", "エラーあり")
            suggestion = error.get("修正アドバイス", "内容を確認してください")

            # 行が存在するかまず確認
            if current_pk_values in pk_to_row:
                row_idx = pk_to_row[current_pk_values]
                
                # 分割された各フィールド名に対して処理を行う
                for col_name in target_cols:
                    clean_col_name = col_name.replace('\n', '').strip()
                    
                    if clean_col_name in col_to_idx:
                        col_idx = col_to_idx[clean_col_name]
                        cell = ws.cell(row=row_idx, column=col_idx)

                        # --- 装飾：背景色とコメントの追加 ---
                        cell.fill = red_fill
                        comment_text = f"【判定理由】: {reason}\n\n【修正アドバイス】: {suggestion}"
                        cell.comment = Comment(comment_text, author="AI Agent")
                        cell.comment.width = 400
                        cell.comment.height = 200
                        success_count += 1
                    else:
                        fail_count += 1
                        log(f"  [MISS] 列名が見つかりません: '{clean_col_name}'")
            else:
                fail_count += 1
                log(f"  [MISS] 行が見つかりません: PK={current_pk_values}")

        # 6. 保存処理
        wb.save(output_path)
        
        # 最終統計を表示
        print(f"--- [FINISH] 処理完了 ---")
        print(f"成功: {success_count} 件, 失敗: {fail_count} 件")
        print(f"保存先: {output_path}")

    except Exception as e:
        print(f"[CRITICAL ERROR] 実行中に予期せぬエラーが発生しました: {e}")
        if DEBUG_MODE:
            traceback.print_exc()

# --- メイン処理 ---
if __name__ == "__main__":
    # 外部のエラーリストファイルを指定
    error_list_file = "error_input.xlsx"

    try:
        # エラーリストを読み込み、辞書形式のリストに変換
        error_df = pd.read_excel(error_list_file, dtype=str)
        error_data_json = error_df.to_dict(orient='records')

        # 主キーとして使用するカラム名
        pks = ["プロジェクトコード", "枝番"]

        # 実行
        mark_and_update_excel_errors("data_dual.xlsx", "data_fixed.xlsx", error_data_json, pks)
    except Exception as e:
        print(f"エラーリストの読み込みに失敗しました: {e}")
        if DEBUG_MODE:
            traceback.print_exc()